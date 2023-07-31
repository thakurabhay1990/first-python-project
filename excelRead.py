# Import Package
import openpyxl

# Load Workbook
workbook = openpyxl.load_workbook("TestData.xlsx")

# # To list all the sheet names in the xlsx file
# print(workbook.sheetnames)
#
# # To check what is an active sheet
# print("Active sheet is " + workbook.active.title)
#
# # Create Object of any sheet on which you want to work
# sheetObject = workbook['Setup']
# print(sheetObject.title)

sheetObject = workbook['Setup']

# To fetch the cell details
# print(sheetObject['A3'].value)
# print(sheetObject['B4'].value)


# cell1=sheetObject.cell(3,1)
# cell1=sheetObject.cell(column=1,row=3)
# print(cell1.value)
# print(cell1.row)
# print(cell1.column)


# How to read complete data from the excel sheet
# Find rows having data
rows = sheetObject.max_row
column = sheetObject.max_column

print("Total Rows are - " + str(rows))
print("Total Column are - " + str(column))

# Approach 1
# for i in range(1,rows+1):
#     for j in range(1,column+1):
#         cell = sheetObject.cell(i,j)
#         print(cell.value)

# Approach 2
# for r in sheetObject['A1':'C4']:
#     for c in r:
#         print(c.value)